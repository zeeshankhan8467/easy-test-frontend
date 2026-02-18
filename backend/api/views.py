from rest_framework import viewsets, status, renderers
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken


class BinaryFileRenderer(renderers.BaseRenderer):
    """Pass bytes through unchanged for file download responses."""
    media_type = 'application/octet-stream'
    format = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data if isinstance(data, bytes) else b''
from django.db.models import Q, Count, Avg, Max, Min, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_GET
from rest_framework_simplejwt.authentication import JWTAuthentication
from datetime import timedelta
import pandas as pd
import json
import logging
from io import BytesIO
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

from .models import (
    Exam, Question, ExamQuestion, Participant,
    ExamParticipant, ExamAttempt, Answer
)
from .serializers import (
    UserSerializer, LoginSerializer, ExamSerializer, ExamCreateUpdateSerializer,
    QuestionSerializer, ParticipantSerializer, ExamParticipantSerializer,
    ExamAttemptSerializer, QuestionAnalysisSerializer, ParticipantResultSerializer,
    ExamReportSerializer, DashboardStatsSerializer, RecentExamSerializer,
    PerformanceDataSerializer, DashboardDataSerializer, LeaderboardEntrySerializer,
    LeaderboardSerializer
)


# Authentication Views
@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        
        user_data = UserSerializer(user).data
        
        return Response({
            'token': str(refresh.access_token),
            'user': user_data
        })
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Exam Views
class ExamViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Exam.objects.filter(created_by=self.request.user).prefetch_related('exam_questions__question')
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return ExamCreateUpdateSerializer
        return ExamSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Allow creating exam without questions (draft mode)
        # Questions can be added later before freezing
        exam = serializer.save()
        return Response(ExamSerializer(exam).data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        exam = self.get_object()
        
        # Check if exam can be edited
        if exam.status == 'frozen':
            return Response(
                {'error': 'Cannot edit a frozen exam'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(exam, data=request.data, partial=kwargs.get('partial', False))
        serializer.is_valid(raise_exception=True)
        
        # Allow removing all questions (draft mode)
        # But validate before freezing
        exam = serializer.save()
        return Response(ExamSerializer(exam).data)

    @action(detail=True, methods=['post'])
    def freeze(self, request, pk=None):
        exam = self.get_object()
        
        # Validate exam can be frozen
        if exam.status == 'frozen':
            return Response(
                {'error': 'Exam is already frozen'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if exam.exam_questions.count() == 0:
            return Response(
                {'error': 'Cannot freeze an exam without questions'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate snapshot
        import hashlib
        import json as json_lib
        
        questions = exam.exam_questions.select_related('question').order_by('order')
        snapshot_data = {
            'exam_id': exam.id,
            'title': exam.title,
            'description': exam.description,
            'duration': exam.duration,
            'revisable': exam.revisable,
            'frozen_at': timezone.now().isoformat(),
            'questions': []
        }
        
        for eq in questions:
            snapshot_data['questions'].append({
                'question_id': eq.question.id,
                'order': eq.order,
                'text': eq.question.text,
                'type': eq.question.type,
                'options': eq.question.options,
                'correct_answer': eq.question.correct_answer,
                'difficulty': eq.question.difficulty,
                'positive_marks': float(eq.positive_marks),
                'negative_marks': float(eq.negative_marks),
                'is_optional': eq.is_optional,
            })
        
        # Generate version/checksum
        snapshot_json = json_lib.dumps(snapshot_data, sort_keys=True)
        snapshot_version = hashlib.md5(snapshot_json.encode()).hexdigest()
        
        # Freeze exam
        exam.status = 'frozen'
        exam.frozen = True
        exam.snapshot_data = snapshot_data
        exam.snapshot_version = snapshot_version
        exam.save()
        
        return Response(ExamSerializer(exam).data)

    @action(detail=True, methods=['get'], url_path='export')
    def export_report_action(self, request, pk=None):
        """Export exam report as Excel or CSV. Query param: format=excel|csv. Uses same access as exam_report (any authenticated user)."""
        try:
            exam = Exam.objects.get(pk=pk)
        except Exam.DoesNotExist:
            return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
        return _build_export_http_response(request, exam)

    @action(detail=False, methods=['get'])
    def available_questions(self, request):
        """Get available questions for selection (not already in exam)"""
        exam_id = request.query_params.get('exam_id')
        difficulty = request.query_params.get('difficulty')
        qtype = request.query_params.get('type')
        search = request.query_params.get('search', '')
        
        queryset = Question.objects.all()
        
        # Filter by difficulty
        if difficulty and difficulty != 'all':
            queryset = queryset.filter(difficulty=difficulty)
        
        # Filter by type
        if qtype and qtype != 'all':
            queryset = queryset.filter(type=qtype)
        
        # Search in text
        if search:
            queryset = queryset.filter(text__icontains=search)
        
        # Exclude questions already in exam (if editing)
        if exam_id:
            try:
                exam = Exam.objects.get(id=exam_id, created_by=request.user)
                existing_question_ids = exam.exam_questions.values_list('question_id', flat=True)
                queryset = queryset.exclude(id__in=existing_question_ids)
            except Exam.DoesNotExist:
                pass
        
        serializer = QuestionSerializer(queryset[:100], many=True)  # Limit to 100
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def snapshot(self, request, pk=None):
        exam = self.get_object()
        
        # Return stored snapshot if frozen, otherwise generate current snapshot
        if exam.snapshot_data:
            snapshot_data = exam.snapshot_data
        else:
            questions = exam.exam_questions.select_related('question').order_by('order')
            snapshot_data = {
                'exam_id': exam.id,
                'title': exam.title,
                'description': exam.description,
                'duration': exam.duration,
                'revisable': exam.revisable,
                'generated_at': timezone.now().isoformat(),
                'questions': []
            }
            
            for eq in questions:
                snapshot_data['questions'].append({
                    'question_id': eq.question.id,
                    'order': eq.order,
                    'text': eq.question.text,
                    'type': eq.question.type,
                    'options': eq.question.options,
                    'correct_answer': eq.question.correct_answer,
                    'difficulty': eq.question.difficulty,
                    'positive_marks': float(eq.positive_marks),
                    'negative_marks': float(eq.negative_marks),
                    'is_optional': eq.is_optional,
                })
        
        # Include version for client validation (only when frozen)
        if exam.snapshot_version:
            snapshot_data['snapshot_version'] = exam.snapshot_version
        response = Response(snapshot_data)
        response['Content-Disposition'] = f'attachment; filename="exam-{exam.id}-snapshot.json"'
        return response

    @action(detail=True, methods=['post'])
    def sync_live_results(self, request, pk=None):
        """
        Submit live clicker responses and attendance from EasyTest Live app.
        Body: {
            "responses": [
                {"participant_id": 1, "question_id": 2, "selected_answer": 0, "answered_at": "ISO8601"},
                {"clicker_id": "123", ...}  // alternative: resolve participant by clicker_id
            ],
            "attendance": [1, 2, 3]  // optional: participant_ids to mark present
        }
        selected_answer: 0-based index (MCQ) or list of indices (multiple_select). Letter "A"=0, "B"=1, etc.
        """
        exam = self.get_object()
        if exam.status not in ('frozen', 'completed'):
            logger.warning('[sync_live_results] Exam %s not frozen/completed, status=%s', pk, exam.status)
            return Response(
                {'error': 'Exam must be frozen to accept live results'},
                status=status.HTTP_400_BAD_REQUEST
            )

        responses_data = request.data.get('responses', [])
        attendance_ids = request.data.get('attendance', [])
        logger.info(
            '[sync_live_results] Exam id=%s: received %d responses, %d attendance',
            pk, len(responses_data), len(attendance_ids)
        )

        # Build snapshot question lookup (question_id -> correct_answer, positive_marks, negative_marks)
        snapshot = exam.snapshot_data or {}
        questions_list = snapshot.get('questions', [])
        if questions_list:
            snapshot_questions = {int(q['question_id']): q for q in questions_list}
        else:
            snapshot_questions = {}
            for eq in exam.exam_questions.select_related('question').order_by('order'):
                snapshot_questions[eq.question_id] = {
                    'correct_answer': eq.question.correct_answer,
                    'positive_marks': float(eq.positive_marks),
                    'negative_marks': float(eq.negative_marks),
                }

        # Resolve participant by id or clicker_id; create one from clicker_id if missing (e.g. deviceId fallback when SDK keySN is empty)
        def get_or_create_participant_for_clicker(participant_id=None, clicker_id=None):
            if participant_id:
                try:
                    p = Participant.objects.get(id=participant_id)
                    ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                    return p
                except Participant.DoesNotExist:
                    pass
            if clicker_id:
                clicker_id_str = str(clicker_id).strip()
                if not clicker_id_str:
                    return None
                try:
                    p = Participant.objects.get(clicker_id=clicker_id_str)
                    if ExamParticipant.objects.filter(exam=exam, participant=p).exists():
                        return p
                    ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                    return p
                except Participant.DoesNotExist:
                    pass
                # When app sends deviceId fallback (d1_timestamp) because SDK keySN is empty, match to participant with clicker_id = number (e.g. "1")
                if clicker_id_str.startswith('d') and '_' in clicker_id_str:
                    num_part = clicker_id_str[1:].split('_')[0]
                    if num_part.isdigit():
                        try:
                            p = Participant.objects.get(clicker_id=num_part)
                            ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                            return p
                        except Participant.DoesNotExist:
                            pass
                # Auto-create participant only if no match (e.g. when SDK sends empty keySN and no participant has that clicker number)
                safe_id = ''.join(c if c.isalnum() or c in '_-' else '_' for c in clicker_id_str)[:50]
                email = f'clicker-{safe_id}@easytest.local'
                p, created = Participant.objects.get_or_create(
                    clicker_id=clicker_id_str,
                    defaults={'name': 'Student', 'email': email}
                )
                if created:
                    logger.info('[sync_live_results] Auto-created participant id=%s clicker_id=%s for exam %s', p.id, clicker_id_str, pk)
                ExamParticipant.objects.get_or_create(exam=exam, participant=p)
                return p
            return None

        # Normalize selected_answer to 0-based index or list
        def normalize_selected(val):
            if val is None:
                return None
            if isinstance(val, list):
                return [int(x) if isinstance(x, int) else (ord(str(x).upper()[0]) - 65) for x in val]
            if isinstance(val, int):
                return val
            s = str(val).strip().upper()
            if s and s[0] in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
                return ord(s[0]) - 65
            try:
                return int(val)
            except (TypeError, ValueError):
                return None

        def is_correct(question_info, selected):
            correct = question_info.get('correct_answer')
            if correct is None:
                return False
            if isinstance(correct, list):
                return sorted(selected if isinstance(selected, list) else [selected]) == sorted(correct)
            return (selected if isinstance(selected, int) else (selected[0] if selected else None)) == correct

        created_attempts = {}
        participant_names = {}  # clicker_id or participant_id -> name for live app display
        answers_created = 0
        skipped_no_participant = 0
        skipped_no_question = 0
        skipped_already_answered = 0
        for item in responses_data:
            participant = get_or_create_participant_for_clicker(
                participant_id=item.get('participant_id'),
                clicker_id=item.get('clicker_id')
            )
            if not participant:
                skipped_no_participant += 1
                continue
            cid = item.get('clicker_id')
            if cid is not None:
                participant_names[str(cid)] = participant.name
            participant_names[str(participant.id)] = participant.name
            question_id = item.get('question_id')
            if question_id is None:
                continue
            question_id = int(question_id)
            if question_id not in snapshot_questions:
                skipped_no_question += 1
                continue
            qinfo = snapshot_questions[question_id]
            selected = normalize_selected(item.get('selected_answer'))
            if selected is None:
                continue

            attempt, _ = ExamAttempt.objects.get_or_create(
                exam=exam,
                participant=participant,
                defaults={'total_questions': len(snapshot_questions)}
            )
            created_attempts[participant.id] = attempt

            # One response per participant per question - do not overwrite
            if Answer.objects.filter(attempt=attempt, question_id=question_id).exists():
                skipped_already_answered += 1
                continue

            correct = is_correct(qinfo, selected)
            pos = float(qinfo.get('positive_marks', 1))
            neg = float(qinfo.get('negative_marks', 0))
            time_taken = 0
            if item.get('answered_at') and attempt.started_at:
                try:
                    raw = item['answered_at']
                    if isinstance(raw, str):
                        from datetime import datetime
                        answered = datetime.fromisoformat(raw.replace('Z', '+00:00'))
                        if timezone.is_naive(answered):
                            answered = timezone.make_aware(answered)
                        time_taken = max(0, int((answered - attempt.started_at).total_seconds()))
                except Exception:
                    pass

            Answer.objects.create(
                attempt=attempt,
                question_id=question_id,
                selected_answer=selected if isinstance(selected, list) else [selected],
                is_correct=correct,
                time_taken=time_taken
            )
            answers_created += 1

        logger.info(
            '[sync_live_results] Exam id=%s: answers_created=%d, attempts_updated=%d, '
            'skipped_no_participant=%d, skipped_no_question=%d, skipped_already_answered=%d',
            pk, answers_created, len(created_attempts),
            skipped_no_participant, skipped_no_question, skipped_already_answered
        )

        # Recalculate attempt totals and submitted_at
        for attempt in created_attempts.values():
            answers = Answer.objects.filter(attempt=attempt)
            total = attempt.total_questions or 0
            correct_count = answers.filter(is_correct=True).count()
            wrong_count = answers.filter(is_correct=False).count()
            unattempted = max(0, total - answers.count())
            total_marks = sum(float(snapshot_questions.get(a.question_id, {}).get('positive_marks', 1)) for a in answers.filter(is_correct=True))
            total_marks -= sum(float(snapshot_questions.get(a.question_id, {}).get('negative_marks', 0)) for a in answers.filter(is_correct=False))
            attempt.correct_answers = correct_count
            attempt.wrong_answers = wrong_count
            attempt.unattempted = unattempted
            attempt.score = total_marks
            attempt.submitted_at = timezone.now()
            if answers.exists():
                last_ans = answers.order_by('-answered_at').first()
                if last_ans and attempt.started_at:
                    attempt.time_taken = max(0, int((last_ans.answered_at - attempt.started_at).total_seconds()))
            attempt.save()

        # Mark attendance for listed participants (at least one response already marks present; this can add late-joiners)
        for pid in attendance_ids:
            try:
                p = Participant.objects.get(id=pid)
                if ExamParticipant.objects.filter(exam=exam, participant=p).exists():
                    ExamAttempt.objects.get_or_create(
                        exam=exam,
                        participant=p,
                        defaults={'total_questions': len(snapshot_questions), 'submitted_at': timezone.now()}
                    )
            except Participant.DoesNotExist:
                pass

        logger.info(
            '[sync_live_results] Exam id=%s: done. synced=%d, attempts_updated=%d',
            pk, answers_created, len(created_attempts)
        )
        return Response({
            'synced': answers_created,
            'attempts_updated': len(created_attempts),
            'received': len(responses_data),
            'skipped_no_participant': skipped_no_participant,
            'skipped_no_question': skipped_no_question,
            'skipped_already_answered': skipped_already_answered,
            'participant_names': participant_names,
        })

    @action(detail=True, methods=['get'], url_path='attendance')
    def attendance(self, request, pk=None):
        """
        GET /api/exams/{id}/attendance/
        Returns attendance for the exam: list of participants with present/absent.
        Present = has an ExamAttempt for this exam (taken attendance or submitted answers).
        """
        exam = self.get_object()
        assigned = ExamParticipant.objects.filter(exam=exam).select_related('participant')
        present_ids = set(
            ExamAttempt.objects.filter(exam=exam).values_list('participant_id', flat=True)
        )
        participants = []
        for ep in assigned:
            p = ep.participant
            participants.append({
                'id': p.id,
                'name': p.name,
                'email': p.email,
                'present': p.id in present_ids,
            })
        present_count = len(present_ids)
        total_count = len(participants)
        return Response({
            'exam_id': exam.id,
            'exam_title': exam.title,
            'participants': participants,
            'present_count': present_count,
            'total_count': total_count,
        })


# Question Views
class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """Generate questions using AI"""
        topic = request.data.get('topic', '').strip()
        count = int(request.data.get('count', 5))
        difficulty = request.data.get('difficulty', 'medium')
        qtype = request.data.get('type', 'mcq')
        
        # Validate inputs
        if not topic:
            return Response(
                {'error': 'Topic is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if count < 1 or count > 20:
            return Response(
                {'error': 'Count must be between 1 and 20'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if difficulty not in ['easy', 'medium', 'hard']:
            return Response(
                {'error': 'Difficulty must be easy, medium, or hard'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if qtype not in ['mcq', 'true_false', 'multiple_select']:
            return Response(
                {'error': 'Type must be mcq, true_false, or multiple_select'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Try AI generation - try Gemini first (free), then OpenAI
            ai_questions = []
            
            # Try Gemini (free tier) first
            try:
                from api.services.ai_generator_gemini import GeminiQuestionGenerator
                generator = GeminiQuestionGenerator()
                ai_questions = generator.generate_questions_safe(topic, count, difficulty, qtype)
            except (ValueError, ImportError) as e:
                # Gemini not available, try OpenAI
                try:
                    from api.services.ai_generator import AIQuestionGenerator
                    generator = AIQuestionGenerator()
                    ai_questions = generator.generate_questions_safe(topic, count, difficulty, qtype)
                except (ValueError, ImportError):
                    # Neither available
                    pass
            
            # If AI generation succeeded and returned questions
            if ai_questions:
                generated_questions = []
                for q_data in ai_questions:
                    question = Question.objects.create(
                        text=q_data['text'],
                        type=qtype,
                        options=q_data['options'],
                        correct_answer=q_data['correct_answer'],
                        difficulty=q_data.get('difficulty', difficulty),
                        tags=q_data.get('tags', [topic]),
                        marks=q_data.get('marks', 1.0)
                    )
                    generated_questions.append(QuestionSerializer(question).data)
                
                return Response(generated_questions, status=status.HTTP_201_CREATED)
            else:
                # Fallback to mock if AI fails
                mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
                return Response(
                    {
                        'questions': mock_data,
                        'warning': 'AI generation unavailable. Sample questions generated. Please check OpenAI account quota or API key configuration.'
                    },
                    status=status.HTTP_201_CREATED
                )
                
        except ValueError as e:
            # API key not set - use mock
            mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
            return Response(
                {
                    'questions': mock_data,
                    'warning': 'OpenAI API key not configured. Sample questions generated.'
                },
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            # Any other error - use mock
            import logging
            logger = logging.getLogger(__name__)
            error_msg = str(e)
            logger.error(f"AI generation error: {error_msg}")
            
            # Check for specific error types
            warning_msg = 'AI generation unavailable. Sample questions generated.'
            if 'quota' in error_msg.lower() or '429' in error_msg:
                warning_msg = 'OpenAI account quota exceeded. Please add credits to your OpenAI account. Sample questions generated.'
            elif 'api key' in error_msg.lower() or '401' in error_msg:
                warning_msg = 'Invalid OpenAI API key. Please check your API key configuration. Sample questions generated.'
            
            mock_data = self._generate_mock_questions(topic, count, difficulty, qtype)
            return Response(
                {
                    'questions': mock_data,
                    'warning': warning_msg
                },
                status=status.HTTP_201_CREATED
            )
    
    def _generate_mock_questions(self, topic: str, count: int, difficulty: str, qtype: str):
        """Fallback mock question generation"""
        generated_questions = []
        for i in range(count):
            question = Question.objects.create(
                text=f"Sample question about {topic} - Question {i+1}",
                type=qtype,
                options=['Option A', 'Option B', 'Option C', 'Option D'] if qtype == 'mcq' else ['True', 'False'],
                correct_answer=0 if qtype == 'mcq' else [0],
                difficulty=difficulty,
                tags=[topic],
                marks=1.0
            )
            generated_questions.append(QuestionSerializer(question).data)
        
        return generated_questions  # Return list directly

    @action(detail=False, methods=['post'], url_path='import')
    def import_questions(self, request):
        """Import questions from CSV or Excel. Required: text, options, correct_answer. Optional: type, difficulty, marks, tags."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            else:
                return Response({'error': 'Unsupported file format. Use CSV or Excel.'}, status=status.HTTP_400_BAD_REQUEST)

            def get_col(df, *candidates):
                for c in candidates:
                    for col in df.columns:
                        if str(col).strip().lower() == c.lower():
                            return col
                return None

            text_col = get_col(df, 'text', 'question', 'question text', 'questions')
            options_col = get_col(df, 'options', 'option', 'choices')
            correct_col = get_col(df, 'correct_answer', 'correct answer', 'correct', 'answer', 'answer_index')
            type_col = get_col(df, 'type', 'question type', 'qtype')
            difficulty_col = get_col(df, 'difficulty')
            marks_col = get_col(df, 'marks', 'mark', 'points')
            tags_col = get_col(df, 'tags', 'tag')

            if text_col is None or options_col is None or correct_col is None:
                return Response(
                    {'error': 'File must have columns: text (or question), options, correct_answer (or correct/answer).'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            imported = 0
            errors = []
            for idx, row in df.iterrows():
                try:
                    raw_text = row.get(text_col, '')
                    text = '' if pd.isna(raw_text) else str(raw_text).strip()
                    if not text:
                        errors.append(f"Row {idx + 2}: Question text is required.")
                        continue

                    raw_opts = row.get(options_col, '')
                    opts_str = '' if pd.isna(raw_opts) else str(raw_opts).strip()
                    if not opts_str:
                        errors.append(f"Row {idx + 2}: Options are required (use pipe | or semicolon ; to separate).")
                        continue
                    options = [o.strip() for o in opts_str.replace(';', '|').split('|') if o.strip()]
                    if len(options) < 2:
                        errors.append(f"Row {idx + 2}: At least 2 options required.")
                        continue

                    raw_correct = row.get(correct_col, '')
                    correct_str = '' if pd.isna(raw_correct) else str(raw_correct).strip()
                    if correct_str == '':
                        errors.append(f"Row {idx + 2}: correct_answer is required (0-based index or comma-separated for multiple select).")
                        continue
                    if ',' in correct_str:
                        correct_answer = [int(x.strip()) for x in correct_str.split(',') if x.strip() != '']
                        if not correct_answer or any(i < 0 or i >= len(options) for i in correct_answer):
                            errors.append(f"Row {idx + 2}: correct_answer indices must be 0 to {len(options) - 1}.")
                            continue
                    else:
                        try:
                            correct_answer = int(correct_str)
                        except ValueError:
                            errors.append(f"Row {idx + 2}: correct_answer must be a number or comma-separated numbers.")
                            continue
                        if correct_answer < 0 or correct_answer >= len(options):
                            errors.append(f"Row {idx + 2}: correct_answer must be 0 to {len(options) - 1}.")
                            continue

                    qtype = 'mcq'
                    if type_col:
                        raw_type = row.get(type_col, '')
                        t = '' if pd.isna(raw_type) else str(raw_type).strip().lower()
                        if t in ('mcq', 'true_false', 'multiple_select'):
                            qtype = t
                        elif t in ('tf', 'true false'):
                            qtype = 'true_false'
                        elif t in ('ms', 'multiple select', 'multi'):
                            qtype = 'multiple_select'

                    diff = 'medium'
                    if difficulty_col:
                        raw_diff = row.get(difficulty_col, '')
                        d = '' if pd.isna(raw_diff) else str(raw_diff).strip().lower()
                        if d in ('easy', 'medium', 'hard'):
                            diff = d

                    marks_val = 1.0
                    if marks_col:
                        raw_marks = row.get(marks_col, '')
                        if pd.notna(raw_marks):
                            try:
                                marks_val = float(raw_marks)
                            except (ValueError, TypeError):
                                pass

                    tags_list = []
                    if tags_col:
                        raw_tags = row.get(tags_col, '')
                        if pd.notna(raw_tags) and str(raw_tags).strip():
                            tags_list = [t.strip() for t in str(raw_tags).split(',') if t.strip()]

                    Question.objects.create(
                        text=text,
                        type=qtype,
                        options=options,
                        correct_answer=correct_answer,
                        difficulty=diff,
                        tags=tags_list,
                        marks=marks_val
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")

            return Response({'imported': imported, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# Participant Views
class ParticipantViewSet(viewsets.ModelViewSet):
    queryset = Participant.objects.all()
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        exam_id = self.request.query_params.get('exam_id')
        user = self.request.user
        if exam_id:
            # Only return participants for exams owned by current user
            try:
                exam = Exam.objects.get(id=exam_id, created_by=user)
            except Exam.DoesNotExist:
                return Participant.objects.none()
            participant_ids = ExamParticipant.objects.filter(exam=exam).values_list('participant_id', flat=True)
            return Participant.objects.filter(id__in=participant_ids)
        # No exam_id (e.g. /participants/ page): return all participants so the list
        # shows every participant and which clicker_id is assigned (avoids "clicker id
        # already assigned" when the assigning participant is not in the filtered list).
        return Participant.objects.all()

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple participants in one request. Each item needs name and clicker_id; email optional."""
        from .serializers import ParticipantBulkCreateSerializer
        ser = ParticipantBulkCreateSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        RESERVED_KEYS = {'name', 'clicker_id', 'email'}
        created = []
        errors = []
        for i, item in enumerate(ser.validated_data['participants']):
            name = (item.get('name') or '').strip()
            clicker_id = (item.get('clicker_id') or '').strip()
            email = (item.get('email') or '').strip() or None
            extra = {k: (v if isinstance(v, str) else str(v)) for k, v in item.items()
                     if k not in RESERVED_KEYS and v is not None and str(v).strip() != ''}
            try:
                participant = Participant.objects.create(
                    name=name,
                    clicker_id=clicker_id,
                    email=email,
                    extra=extra
                )
                created.append(ParticipantSerializer(participant).data)
            except Exception as e:
                errors.append(f"Row {i + 1}: {str(e)}")
        return Response(
            {'created': len(created), 'participants': created, 'errors': errors},
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['post'], url_path='import')
    def import_participants(self, request):
        file = request.FILES.get('file')
        exam_id = request.data.get('exam_id')
        
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            else:
                return Response({'error': 'Unsupported file format'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Flexible column mapping: name and clicker_id required; email optional
            def get_col(df, *candidates):
                for c in candidates:
                    for col in df.columns:
                        if str(col).strip().lower() == c.lower():
                            return col
                return None
            name_col = get_col(df, 'name', 'names', 'participant', 'participant name')
            clicker_col = get_col(df, 'clicker_id', 'clicker id', 'clickerid', 'clicker')
            email_col = get_col(df, 'email', 'e-mail', 'email address')
            
            if name_col is None or clicker_col is None:
                return Response(
                    {'error': 'File must have at least "name" and "clicker_id" (or "clicker id") columns.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            reserved = {name_col, clicker_col, email_col} if email_col else {name_col, clicker_col}
            imported = 0
            errors = []
            for idx, row in df.iterrows():
                try:
                    raw_name = row.get(name_col, '')
                    raw_clicker = row.get(clicker_col, '')
                    name = '' if pd.isna(raw_name) else str(raw_name).strip()
                    clicker_id = '' if pd.isna(raw_clicker) else str(raw_clicker).strip()
                    email = None
                    if email_col:
                        raw_email = row.get(email_col, '')
                        email = None if pd.isna(raw_email) else (str(raw_email).strip() or None)
                    extra = {}
                    for col in df.columns:
                        if col not in reserved:
                            val = row.get(col)
                            if pd.notna(val) and val is not None and str(val).strip() not in ('', 'nan'):
                                extra[str(col).strip()] = str(val).strip()
                    
                    if not name or not clicker_id:
                        errors.append(f"Row {idx + 2}: Name and Clicker ID are required.")
                        continue
                    
                    participant, created = Participant.objects.update_or_create(
                        clicker_id=clicker_id,
                        defaults={'name': name, 'email': email, 'extra': extra}
                    )
                    
                    if exam_id:
                        try:
                            exam = Exam.objects.get(id=exam_id, created_by=request.user)
                            ExamParticipant.objects.get_or_create(exam=exam, participant=participant)
                        except Exam.DoesNotExist:
                            pass
                    
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")
            
            return Response({'imported': imported, 'errors': errors})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def assign_clicker(self, request, pk=None):
        participant = self.get_object()
        clicker_id = request.data.get('clicker_id')
        
        if not clicker_id:
            return Response({'error': 'clicker_id required'}, status=status.HTTP_400_BAD_REQUEST)
        
        participant.clicker_id = clicker_id
        participant.save()
        
        return Response(ParticipantSerializer(participant).data)


# Reports Views
def _get_exam_report_data(exam):
    """Build report data dict for an exam (no request needed). Used by exam_report and export."""
    attempts = ExamAttempt.objects.filter(exam=exam)
    total_participants = attempts.count()
    if total_participants == 0:
        return {
            'exam_id': exam.id,
            'exam_title': exam.title,
            'total_participants': 0,
            'average_score': 0,
            'highest_score': 0,
            'lowest_score': 0,
            'question_analysis': [],
            'participant_results': []
        }
    avg_score = attempts.aggregate(avg=Avg('score'))['avg'] or 0
    highest = attempts.aggregate(max=Max('score'))['max'] or 0
    lowest = attempts.aggregate(min=Min('score'))['min'] or 0
    questions = exam.exam_questions.select_related('question').order_by('order')
    question_analysis = []
    for eq in questions:
        question = eq.question
        answers = Answer.objects.filter(attempt__exam=exam, question=question)
        total_attempts = answers.count()
        correct_attempts = answers.filter(is_correct=True).count()
        accuracy = (correct_attempts / total_attempts * 100) if total_attempts > 0 else 0
        avg_time = answers.aggregate(avg=Avg('time_taken'))['avg'] or 0
        question_analysis.append({
            'question_id': question.id,
            'question_text': question.text,
            'total_attempts': total_attempts,
            'correct_attempts': correct_attempts,
            'accuracy': accuracy,
            'average_time': avg_time
        })
    participant_results = []
    for attempt in attempts.order_by('-score', 'time_taken'):
        participant_results.append({
            'participant_id': attempt.participant.id,
            'participant_name': attempt.participant.name,
            'score': float(attempt.score),
            'total_questions': attempt.total_questions,
            'correct_answers': attempt.correct_answers,
            'wrong_answers': attempt.wrong_answers,
            'unattempted': attempt.unattempted,
            'percentage': float(attempt.percentage),
            'rank': 0
        })
    for idx, result in enumerate(participant_results, 1):
        result['rank'] = idx
    return {
        'exam_id': exam.id,
        'exam_title': exam.title,
        'total_participants': total_participants,
        'average_score': float(avg_score),
        'highest_score': float(highest),
        'lowest_score': float(lowest),
        'question_analysis': question_analysis,
        'participant_results': participant_results
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exam_report(request, exam_id):
    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
    report_data = _get_exam_report_data(exam)
    serializer = ExamReportSerializer(report_data)
    return Response(serializer.data)


def _write_results_by_participants_detail_sheet(workbook, exam):
    """Write 'Results by Participants (Detail)' sheet: exam metadata + per-participant blocks with Question, Option, Type Correct Answer, Speed, Score."""
    # Excel sheet names cannot contain [ ] \ / * ? :
    sheet_name = 'Results by Participants (Detail)'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    attempts = list(ExamAttempt.objects.filter(exam=exam).order_by('-score', 'time_taken').select_related('participant'))
    answer_map = {}  # (attempt_id, question_id) -> Answer
    for a in Answer.objects.filter(attempt__exam=exam).select_related('attempt', 'question'):
        answer_map[(a.attempt_id, a.question_id)] = a

    row_num = 1
    # Title
    ws.cell(row=row_num, column=1, value='Results by Participants [Detail]')  # display title (brackets OK in cell)
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    row_num += 1
    # Date and Questions Count
    ws.cell(row=row_num, column=1, value=f'Date: {timezone.now().strftime("%m/%d/%Y %H:%M")}')
    row_num += 1
    ws.cell(row=row_num, column=1, value=f'Questions Count: {len(exam_questions)}')
    row_num += 2

    for attempt in attempts:
        participant = attempt.participant
        keypad = getattr(participant, 'clicker_id', None) or participant.name or str(participant.id)
        # Keypad No.
        ws.cell(row=row_num, column=1, value=f'Keypad No. {keypad}')
        ws.cell(row=row_num, column=1).font = bold_font
        row_num += 1
        # Correct Co Score / Correct Rs Ranking
        ws.cell(row=row_num, column=1, value=f'Correct Co Score: {float(attempt.score)}')
        ws.cell(row=row_num, column=2, value=f'Correct Rs Ranking: {attempt.correct_answers}/{attempt.total_questions}')
        row_num += 1
        # Table header
        headers = ['Question', 'Option', 'Response Date', 'Type Correct Answer', 'Speed', 'Score']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.fill = red_header_fill
            c.font = white_font
        row_num += 1
        # One row per exam question
        for eq in exam_questions:
            q = eq.question
            options_text = '\n'.join([f'{i+1}. {opt}' for i, opt in enumerate(q.options or [])])
            answer = answer_map.get((attempt.id, q.id))
            if answer:
                sel = answer.selected_answer
                if isinstance(sel, list):
                    type_correct = ', '.join(str(s + 1) for s in sel) + ' Choice'
                else:
                    type_correct = f'{int(sel) + 1} Choice' if sel is not None else ''
                response_date = answer.answered_at.strftime('%m/%d/%Y %H:%M') if answer.answered_at else ''
                speed = answer.time_taken  # seconds; can show as decimal e.g. round(answer.time_taken, 2)
                score = float(eq.positive_marks) if answer.is_correct else -float(eq.negative_marks)
            else:
                type_correct = ''
                response_date = ''
                speed = ''
                score = 0
            ws.cell(row=row_num, column=1, value=q.text)
            ws.cell(row=row_num, column=2, value=options_text)
            ws.cell(row=row_num, column=3, value=response_date)
            ws.cell(row=row_num, column=4, value=type_correct)
            ws.cell(row=row_num, column=5, value=speed)
            ws.cell(row=row_num, column=6, value=score)
            row_num += 1
        row_num += 1  # blank row between participants

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8


def _sanitize_sheet_name(name):
    """Excel sheet names: max 31 chars, no \\ / * ? : [ ]"""
    for c in '\\/*?:[]':
        name = name.replace(c, '_')
    return (name or 'Sheet')[:31]


def _write_results_by_participants_individual(workbook, exam):
    """One sheet per participant (sheet name = keypad id). Same layout as reference: title, date, questions count, metadata, Keypad summary, then Question|Option|Response|Slide Type|Correct Answer|Speed|Score."""
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    attempts = list(ExamAttempt.objects.filter(exam=exam).order_by('-score', 'time_taken').select_related('participant'))
    answer_map = {}
    for a in Answer.objects.filter(attempt__exam=exam).select_related('attempt', 'question'):
        answer_map[(a.attempt_id, a.question_id)] = a

    n_questions = len(exam_questions)
    report_date = timezone.now().strftime('%m/%d/%Y %H:%M')

    for attempt in attempts:
        participant = attempt.participant
        keypad = getattr(participant, 'clicker_id', None) or str(participant.id)
        sheet_name = _sanitize_sheet_name(str(keypad))
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)

        row = 1
        ws.cell(row=row, column=1, value='Results by Participants(Detail)')
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=report_date)
        row += 1
        ws.cell(row=row, column=1, value=f'Questions Count: {n_questions}')
        row += 2
        # Metadata placeholder rows (Sl No., Exam Unique Id, Exam Name, Theme, Faculty, Subject, Batch)
        for label in ['Sl No.', 'Exam Unique Id', 'Exam Name', 'Theme', 'Faculty', 'Subject', 'Batch']:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value='d')
            row += 1
        row += 2
        # Keypad summary
        correct_rate = (attempt.correct_answers / attempt.total_questions * 100) if attempt.total_questions else 0
        ws.cell(row=row, column=1, value=f'Keypad No. {keypad}')
        ws.cell(row=row, column=2, value=f'Correct Count: {attempt.correct_answers}')
        ws.cell(row=row, column=3, value=f'Score: {int(attempt.score)}')
        row += 1
        ws.cell(row=row, column=2, value=f'Correct Rate: {correct_rate:.2f}%')
        ws.cell(row=row, column=3, value=f'Ranking: {attempt.correct_answers}/{attempt.total_questions}')
        row += 1
        # Table header
        headers = ['Question', 'Option', 'Response', 'Slide Type', 'Correct Answer', 'Speed', 'Score']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = red_header_fill
            c.font = white_font
        row += 1
        # One row per question (match reference: Response = 1-based index, Slide Type = 'Choice', Correct Answer = 1-based correct index, Speed = decimal)
        for order_idx, eq in enumerate(exam_questions, start=1):
            q = eq.question
            options_text = '\n'.join([f'{i+1}. {opt}' for i, opt in enumerate(q.options or [])])
            answer = answer_map.get((attempt.id, q.id))
            correct_idx = q.correct_answer
            if isinstance(correct_idx, list):
                correct_display = correct_idx[0] + 1 if correct_idx else ''
            else:
                correct_display = int(correct_idx) + 1 if correct_idx is not None else ''
            if answer:
                sel = answer.selected_answer
                if isinstance(sel, list):
                    response_val = (sel[0] + 1) if sel else ''
                else:
                    response_val = int(sel) + 1 if sel is not None else ''
                speed = round(answer.time_taken, 2) if answer.time_taken is not None else ''
                score = int(eq.positive_marks) if answer.is_correct else -int(eq.negative_marks)
            else:
                response_val = ''
                speed = ''
                score = 0
            ws.cell(row=row, column=1, value=f'{order_idx}. {q.text}')
            ws.cell(row=row, column=2, value=options_text)
            ws.cell(row=row, column=3, value=response_val)
            ws.cell(row=row, column=4, value='Choice')
            ws.cell(row=row, column=5, value=correct_display)
            ws.cell(row=row, column=6, value=speed)
            ws.cell(row=row, column=7, value=score)
            row += 1

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8

    # Remove default empty sheet if present (we didn't write any pandas sheet first)
    for default in ('Sheet', 'Sheet1'):
        if default in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook[default]
            break


def _write_results_by_questions_sheet(workbook, exam):
    """One sheet 'Results by Questions': per-question blocks with Option, Voted, Percentage (like reference)."""
    sheet_name = 'Results by Questions'
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name, 0)
    red_header_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    white_font = Font(color='FFFFFF', bold=True)

    exam_questions = list(exam.exam_questions.select_related('question').order_by('order'))
    answers_by_question = {}  # question_id -> list of selected_answer (int or list)
    for a in Answer.objects.filter(attempt__exam=exam).values_list('question_id', 'selected_answer'):
        qid, sel = a[0], a[1]
        answers_by_question.setdefault(qid, []).append(sel)


    row = 1
    ws.cell(row=row, column=1, value='Results by Questions')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=exam.title or '')
    row += 1
    ws.cell(row=row, column=1, value=timezone.now().strftime('%m/%d/%Y %I:%M:%S %p'))
    row += 2
    for label in ['Sl No.', 'Exam Unique Id', 'Exam Name', 'Theme', 'Faculty', 'Subject', 'Batch']:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value='d')
        row += 1
    row += 1

    for order_idx, eq in enumerate(exam_questions, start=1):
        q = eq.question
        options = q.options or []
        correct_rate = 0.0
        total_answers = 0
        option_votes = [0] * len(options)  # vote count per option index

        selections = answers_by_question.get(q.id) or []
        total_answers = len(selections)
        for sel in selections:
            if isinstance(sel, list):
                for i in sel:
                    if 0 <= i < len(option_votes):
                        option_votes[i] += 1
            else:
                if 0 <= sel < len(option_votes):
                    option_votes[sel] += 1
        if total_answers > 0:
            correct_count = sum(1 for s in selections if _answer_is_correct(s, q.correct_answer))
            correct_rate = round(correct_count / total_answers * 100, 2)

        ws.cell(row=row, column=1, value=f'{order_idx}. {q.text}')
        ws.cell(row=row, column=2, value='Slide Type: Choice')
        ws.cell(row=row, column=3, value=f'Correct Rate: {correct_rate:.2f}%')
        row += 1
        ws.cell(row=row, column=1, value='Option')
        ws.cell(row=row, column=2, value='Voted')
        ws.cell(row=row, column=3, value='Percentage')
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = red_header_fill
            ws.cell(row=row, column=c).font = white_font
        row += 1

        for i, opt in enumerate(options):
            letter = chr(ord('A') + i) if i < 26 else 'X'
            opt_label = f'{i+1}/{letter}. {opt}'
            vote_count = option_votes[i] if i < len(option_votes) else 0
            pct = round(vote_count / total_answers, 4) if total_answers else 0
            ws.cell(row=row, column=1, value=opt_label)
            ws.cell(row=row, column=2, value=vote_count)
            ws.cell(row=row, column=3, value=pct)
            row += 1
        ws.cell(row=row, column=1, value='Voted')
        ws.cell(row=row, column=2, value=total_answers)
        ws.cell(row=row, column=3, value=1)
        row += 2

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14


def _answer_is_correct(selected, correct_answer):
    """Return True if selected answer matches correct_answer (int or list)."""
    if isinstance(correct_answer, list):
        return isinstance(selected, list) and set(selected) == set(correct_answer)
    return selected == correct_answer


def _build_export_http_response(request, exam):
    """Build Excel or CSV download for exam report. Query param: format=excel|csv (case-insensitive, default excel). Returns DRF Response."""
    raw = (request.query_params.get('format') or 'excel').strip().lower()
    format_type = 'excel' if raw in ('excel', 'xlsx', '') else 'csv'
    report_data = _get_exam_report_data(exam)
    df = pd.DataFrame(report_data['participant_results'])
    output = BytesIO()
    if format_type == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Participant Results', index=False)
            qa_df = pd.DataFrame(report_data['question_analysis'])
            qa_df.to_excel(writer, sheet_name='Question Analysis', index=False)
            _write_results_by_participants_detail_sheet(writer.book, exam)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = f'report-{exam.id}.xlsx'
    else:
        df.to_csv(output, index=False)
        content_type = 'text/csv'
        filename = f'report-{exam.id}.csv'
    output.seek(0)
    return Response(
        output.getvalue(),
        content_type=content_type,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_report(request, exam_id):
    """Export report: GET /api/report-export/<exam_id>/?format=excel|csv. Returns file download as DRF Response."""
    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response(
            {'error': 'Exam not found', 'exam_id': exam_id, 'message': 'No exam with this id. Run: python manage.py seed_dummy_data'},
            status=status.HTTP_404_NOT_FOUND
        )
    return _build_export_http_response(request, exam)


# Use binary renderer so file bytes are not JSON-encoded
export_report.renderer_classes = [BinaryFileRenderer]


def _build_export_file_response(format_type, exam, layout=None):
    """Build Excel or CSV as Django HttpResponse. layout='individual' => one sheet per participant; layout='questions' => Results by Questions."""
    report_data = _get_exam_report_data(exam)
    df = pd.DataFrame(report_data['participant_results'])
    output = BytesIO()
    if format_type == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            layout_l = (layout or '').strip().lower()
            if layout_l == 'individual':
                _write_results_by_participants_individual(writer.book, exam)
                filename = f'report-{exam.id}-individual.xlsx'
            elif layout_l == 'questions':
                _write_results_by_questions_sheet(writer.book, exam)
                # Remove default sheet if pandas created one
                for default in ('Sheet', 'Sheet1'):
                    if default in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
                        del writer.book[default]
                        break
                filename = f'report-{exam.id}-questions.xlsx'
            else:
                df.to_excel(writer, sheet_name='Participant Results', index=False)
                qa_df = pd.DataFrame(report_data['question_analysis'])
                qa_df.to_excel(writer, sheet_name='Question Analysis', index=False)
                _write_results_by_participants_detail_sheet(writer.book, exam)
                filename = f'report-{exam.id}.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:
        df.to_csv(output, index=False)
        content_type = 'text/csv'
        filename = f'report-{exam.id}.csv'
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_GET
def export_report_http(request, exam_id):
    """Plain Django view for report export. Uses HttpRequest and returns HttpResponse to avoid DRF assertion."""
    auth = JWTAuthentication()
    try:
        user_auth = auth.authenticate(request)
    except Exception:
        user_auth = None
    if user_auth is None:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)
    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return JsonResponse(
            {'error': 'Exam not found', 'exam_id': exam_id, 'message': 'No exam with this id. Run: python manage.py seed_dummy_data'},
            status=404
        )
    raw = (request.GET.get('format') or 'excel').strip().lower()
    format_type = 'excel' if raw in ('excel', 'xlsx', '') else 'csv'
    layout = (request.GET.get('layout') or '').strip().lower()  # 'individual' => one sheet per participant
    return _build_export_file_response(format_type, exam, layout=layout)


# Dashboard Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    user = request.user
    exams = Exam.objects.filter(created_by=user)
    
    # Stats
    total_exams = exams.count()
    total_participants = Participant.objects.count()
    
    # Calculate average score from all attempts
    all_attempts = ExamAttempt.objects.filter(exam__created_by=user).select_related('exam')
    if all_attempts.exists():
        total_percentage = 0
        count = 0
        for attempt in all_attempts:
            if attempt.total_questions > 0:
                percentage = (float(attempt.score) / (attempt.total_questions * float(attempt.exam.positive_marking))) * 100
                total_percentage += percentage
                count += 1
        avg_score = total_percentage / count if count > 0 else 0
    else:
        avg_score = 0
    
    # Attendance rate (participants who attempted / total participants)
    total_assigned = ExamParticipant.objects.filter(exam__created_by=user).count()
    attempted = all_attempts.count()
    attendance_rate = (attempted / total_assigned * 100) if total_assigned > 0 else 0
    
    stats = {
        'total_exams': total_exams,
        'total_participants': total_participants,
        'average_score': float(avg_score),
        'attendance_rate': float(attendance_rate)
    }
    
    # Recent exams
    recent_exams = []
    for exam in exams[:5]:
        attempts = ExamAttempt.objects.filter(exam=exam).select_related('exam')
        participant_count = attempts.count()
        if attempts.exists():
            total_percentage = 0
            count = 0
            for attempt in attempts:
                if attempt.total_questions > 0:
                    percentage = (float(attempt.score) / (attempt.total_questions * float(exam.positive_marking))) * 100
                    total_percentage += percentage
                    count += 1
            avg_exam_score = total_percentage / count if count > 0 else 0
        else:
            avg_exam_score = 0
        
        recent_exams.append({
            'id': exam.id,
            'title': exam.title,
            'created_at': exam.created_at,
            'participant_count': participant_count,
            'average_score': float(avg_exam_score)
        })
    
    # Performance data (last 7 days)
    performance_data = []
    for i in range(6, -1, -1):
        date = timezone.now() - timedelta(days=i)
        date_str = date.strftime('%Y-%m-%d')
        
        day_attempts = ExamAttempt.objects.filter(
            exam__created_by=user,
            submitted_at__date=date.date()
        ).select_related('exam')
        
        if day_attempts.exists():
            total_percentage = 0
            count = 0
            for attempt in day_attempts:
                if attempt.total_questions > 0:
                    percentage = (float(attempt.score) / (attempt.total_questions * float(attempt.exam.positive_marking))) * 100
                    total_percentage += percentage
                    count += 1
            avg_day_score = total_percentage / count if count > 0 else 0
        else:
            avg_day_score = 0
        participant_count = day_attempts.values('participant').distinct().count()
        
        performance_data.append({
            'date': date_str,
            'score': float(avg_day_score),
            'participants': participant_count
        })
    
    dashboard_data = {
        'stats': stats,
        'recent_exams': recent_exams,
        'performance_data': performance_data
    }
    
    serializer = DashboardDataSerializer(dashboard_data)
    return Response(serializer.data)


# Leaderboard Views
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def leaderboard(request, exam_id):
    try:
        exam = Exam.objects.get(id=exam_id)
    except Exam.DoesNotExist:
        return Response({'error': 'Exam not found'}, status=status.HTTP_404_NOT_FOUND)
    
    attempts = ExamAttempt.objects.filter(exam=exam).order_by('-score', 'time_taken')
    
    entries = []
    for rank, attempt in enumerate(attempts, 1):
        entries.append({
            'rank': rank,
            'participant_id': attempt.participant.id,
            'participant_name': attempt.participant.name,
            'score': float(attempt.score),
            'percentage': float(attempt.percentage),
            'total_questions': attempt.total_questions,
            'correct_answers': attempt.correct_answers,
            'time_taken': attempt.time_taken
        })
    
    leaderboard_data = {
        'exam_id': exam.id,
        'exam_title': exam.title,
        'entries': entries,
        'generated_at': timezone.now()
    }
    
    serializer = LeaderboardSerializer(leaderboard_data)
    return Response(serializer.data)
